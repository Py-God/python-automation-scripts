#! python3
#  The program should open a spreadsheet and write the cells of column A into one text file, the cells
# of column B into another text file, and so on
import openpyxl, os

wb = openpyxl.load_workbook('txtToSpreadsheet.xlsx')
sheet = wb.active
b = 1
for i in range(1, sheet.max_column+1):
    content = []
    for row in range(1, sheet.max_row):
        value = sheet.cell(row=row, column=b).value
        if value == None:
            continue
        else:
            content.append(value)

    file = open('C:\\Users\\user\\Desktop\\PYTHON\\textFiles\\textFile%s.txt' % b, 'w')
    cont = '\n'.join(content)
    file.write(cont)
    file.close()
    b+=1
