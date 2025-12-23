#! python3
# a that takes two integers and a filename string as command line arguments. Let’s call the first integer N and the second 
# integer M. Starting at row N, the program should insert M blank rows into the spreadsheet.
import openpyxl, sys
from openpyxl.utils import get_column_letter

file = 'C:\\Users\\user\\Desktop\\PYTHON\\lands.xlsx'
if len(sys.argv) < 4:
    print('The command line takes 3 arguments.')
else:
    print('saving the command line arguements...')
    startRow = int(sys.argv[1])
    blankRowNum = int(sys.argv[2])
    filename = sys.argv[3]

    print('Opening lands.xlsx...')
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    print('Opening %s.xlsx' % filename)
    nWb = openpyxl.Workbook()
    sheet1 = nWb.active

    print('''Appending lands.xlsx values to %s
and adding %s blank rows''' % (blankRowNum, filename))
    for row in range(1, startRow):
        for col in range(1, sheet.max_column+1):
            value = sheet.cell(row=row, column=col).value
            sheet1.cell(row=row, column=col).value = value

    for row in range(startRow, startRow+blankRowNum):
        for column in range(1, sheet.max_column):
            sheet1.cell(row=row, column=col).value = ' '

    for row in range(startRow, sheet.max_row+1):
        for col in range(1, sheet.max_column+1):
            for y in range(startRow+blankRowNum, (startRow+blankRowNum)+sheet.max_row+1-startRow):
                value = sheet.cell(row=row, column=col).value
                sheet1.cell(row=y, column=col).value = value

print('Saving %s in your directory...' % filename)

nWb.save('C:\\Users\\user\\Desktop\\PYTHON\\' + filename)

