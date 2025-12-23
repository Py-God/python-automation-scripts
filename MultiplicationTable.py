#! python3
# Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N
# multiplication table in an Excel spreadsheet.
import openpyxl
from openpyxl.styles import Font

def multiplicationTable(number):
    wb = openpyxl.Workbook()
    sheet = wb.active
    boldFont = Font(bold=True)
    for row in range(1, number+1):
        sheet.cell(row=row+1, column=1).value = row
        sheet.cell(row=row+1, column=1).font = boldFont
    for col in range(1, number+1):
        sheet.cell(row=1,column=col+1).value = col
        sheet.cell(row=1, column=col+1).font = boldFont
    for row in range(1, number+1):
        for col in range(number + 1):
            sheet.cell(row=row+1, column=col+1).value = row*col

    wb.save('multiplicationTable.xlsx')

num = int(input('Enter the number: '))
multiplicationTable(num)
