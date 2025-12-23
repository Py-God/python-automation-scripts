#! python3
#  a program to invert the row and column of the cells in the spreadsheet. For example, the value at row 5,
# column 3 will be at row 3, column 5 (and viceversa).
import openpyxl

wb = openpyxl.load_workbook('lands.xlsx')
sheet = wb.active
nwb = openpyxl.Workbook()
sheet1 = nwb.active
for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        y = sheet.cell(row=row, column=col).value
        sheet1.cell(row=col, column=row).value = y
nwb.save('landssss.xlsx')