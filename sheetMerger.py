#! python3
# Read data from one spreadsheet and write it to parts of other spreadsheets.
import openpyxl

print('Opening both spreadsheets...')
exampleWb = openpyxl.load_workbook('example.xlsx')
landWb = openpyxl.load_workbook('lands.xlsx')
exampleSheet = exampleWb['Sheet1']
landSheet = landWb['Sheet1']
print('Getting values from lands.xlsx...')
for row in range(1, landSheet.max_row + 1):
    landName = landSheet['A' + str(row)].value
    landLenght = landSheet['B' + str(row)].value
    landBreath = landSheet['C' + str(row)].value
    print('Appending land.xlsx values to example.xlsx...')
    for cells in exampleSheet['A8':'A10']:
        for cell in cells:
            cell.value = landName
    for cells in exampleSheet['B8':'B10']:
        for cell in cells:
            cell.value = landLenght
    for cells in exampleSheet['C8':'C10']:
        for cell in cells:
            cell.value = landBreath
print('Saving to a new file: newExample.xlsx in your directory...')
exampleWb.save('newExample.xlsx')
